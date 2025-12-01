import tkinter as tk
from tkinter import ttk, messagebox
from db import connect_mysql
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side

def open_invoice_window(root):
    win = tk.Toplevel(root)
    win.title("Lập hóa đơn")
    win.geometry("950x680")

    # ================================
    # TẢI DỮ LIỆU
    # ================================
    def load_staff():
        cn = connect_mysql()
        cur = cn.cursor()
        cur.execute("SELECT id_nv, ho_ten, sdt, dia_chi FROM nhanvien")
        data = cur.fetchall()
        cn.close()
        return data

    def load_customer():
        cn = connect_mysql()
        cur = cn.cursor()
        cur.execute("SELECT id_khachhang, ho_ten, sdt, dia_chi FROM khachhang")
        data = cur.fetchall()
        cn.close()
        return data

    def load_product():
        cn = connect_mysql()
        cur = cn.cursor()
        cur.execute("SELECT id_sanpham, ten_sanpham, gia FROM sanpham")
        data = cur.fetchall()
        cn.close()
        return data

    staff_list = load_staff()
    customer_list = load_customer()
    product_list = load_product()

    # ===========================================================
    #           THÔNG TIN HÓA ĐƠN
    # ===========================================================
    frame_info = tk.LabelFrame(win, text="Thông tin hóa đơn", padx=10, pady=10)
    frame_info.pack(fill="x", padx=10, pady=10)
    frame_info.columnconfigure(1, weight=1)
    frame_info.columnconfigure(3, weight=1)
    frame_info.columnconfigure(5, weight=1)

    # -------------------- NHÂN VIÊN --------------------
    tk.Label(frame_info, text="Nhân viên:").grid(row=0, column=0, sticky="w", pady=4)
    cbo_nv = ttk.Combobox(frame_info, values=[f"{r[0]} - {r[1]}" for r in staff_list])
    cbo_nv.grid(row=0, column=1, sticky="ew", padx=5, pady=4)

    tk.Label(frame_info, text="Tên NV:").grid(row=1, column=0, sticky="w")
    ten_nv = tk.Entry(frame_info); ten_nv.grid(row=1, column=1, sticky="ew", padx=5)

    tk.Label(frame_info, text="SĐT NV:").grid(row=1, column=2, sticky="w")
    sdt_nv = tk.Entry(frame_info); sdt_nv.grid(row=1, column=3, sticky="ew", padx=5)

    tk.Label(frame_info, text="Địa chỉ NV:").grid(row=1, column=4, sticky="w")
    dc_nv = tk.Entry(frame_info); dc_nv.grid(row=1, column=5, sticky="ew", padx=5)

    def on_staff_select(e):
        if not cbo_nv.get():
            return
        id_sel = int(cbo_nv.get().split(" - ")[0])
        for r in staff_list:
            if r[0] == id_sel:
                ten_nv.config(state="normal")
                sdt_nv.config(state="normal")
                dc_nv.config(state="normal")

                ten_nv.delete(0, tk.END); ten_nv.insert(0, r[1])
                sdt_nv.delete(0, tk.END); sdt_nv.insert(0, r[2])
                dc_nv.delete(0, tk.END); dc_nv.insert(0, r[3])

                ten_nv.config(state="disabled")
                sdt_nv.config(state="disabled")
                dc_nv.config(state="disabled")
                break

    cbo_nv.bind("<<ComboboxSelected>>", on_staff_select)

    # -------------------- KHÁCH HÀNG --------------------
    tk.Label(frame_info, text="Khách hàng:").grid(row=2, column=0, sticky="w", pady=4)
    cbo_kh = ttk.Combobox(frame_info,
                          values=[f"{r[0]} - {r[1]}" for r in customer_list])
    cbo_kh.grid(row=2, column=1, sticky="ew", padx=5, pady=4)

    tk.Label(frame_info, text="Tên KH:").grid(row=3, column=0, sticky="w")
    ten_kh = tk.Entry(frame_info); ten_kh.grid(row=3, column=1, sticky="ew", padx=5)

    tk.Label(frame_info, text="SĐT KH:").grid(row=3, column=2, sticky="w")
    sdt_kh = tk.Entry(frame_info); sdt_kh.grid(row=3, column=3, sticky="ew", padx=5)

    tk.Label(frame_info, text="Địa chỉ KH:").grid(row=3, column=4, sticky="w")
    dc_kh = tk.Entry(frame_info); dc_kh.grid(row=3, column=5, sticky="ew", padx=5)

    def on_customer_select(e):
        if not cbo_kh.get():
            return
        id_sel = int(cbo_kh.get().split(" - ")[0])
        for r in customer_list:
            if r[0] == id_sel:
                ten_kh.config(state="normal")
                sdt_kh.config(state="normal")
                dc_kh.config(state="normal")

                ten_kh.delete(0, tk.END); ten_kh.insert(0, r[1])
                sdt_kh.delete(0, tk.END); sdt_kh.insert(0, r[2])
                dc_kh.delete(0, tk.END); dc_kh.insert(0, r[3])

                ten_kh.config(state="disabled")
                sdt_kh.config(state="disabled")
                dc_kh.config(state="disabled")
                break

    cbo_kh.bind("<<ComboboxSelected>>", on_customer_select)

    # -------------------- NGÀY LẬP --------------------
    tk.Label(frame_info, text="Ngày lập:").grid(row=4, column=0, sticky="w")
    ngay = tk.Entry(frame_info)
    today = datetime.now().strftime("%Y-%m-%d")
    ngay.insert(0, today)
    ngay.config(state="disabled")
    ngay.grid(row=4, column=1, sticky="w", padx=5)

    # ===========================================================
    #                     SẢN PHẨM
    # ===========================================================
    frame_sp = tk.LabelFrame(win, text="Sản phẩm", padx=10, pady=10)
    frame_sp.pack(fill="x", padx=10, pady=10)

    frame_sp.columnconfigure(1, weight=1)
    frame_sp.columnconfigure(3, weight=1)
    frame_sp.columnconfigure(5, weight=1)

    tk.Label(frame_sp, text="Chọn sản phẩm:").grid(row=0, column=0, sticky="w")
    cbo_sp = ttk.Combobox(frame_sp, values=[f"{r[0]} - {r[1]}" for r in product_list])
    cbo_sp.grid(row=0, column=1, sticky="ew", padx=5)

    tk.Label(frame_sp, text="Giá:").grid(row=0, column=2, sticky="w")
    gia_sp = tk.Entry(frame_sp, state="disabled")
    gia_sp.grid(row=0, column=3, sticky="ew", padx=5)

    tk.Label(frame_sp, text="Số lượng:").grid(row=0, column=4, sticky="w")
    sl = tk.Entry(frame_sp); sl.insert(0, "1")
    sl.grid(row=0, column=5, sticky="ew", padx=5)

    def on_product_select(e):
        if not cbo_sp.get():
            return
        id_sel = int(cbo_sp.get().split(" - ")[0])
        for r in product_list:
            if r[0] == id_sel:
                gia_sp.config(state="normal")
                gia_sp.delete(0, tk.END)
                gia_sp.insert(0, r[2])
                gia_sp.config(state="disabled")
                break

    cbo_sp.bind("<<ComboboxSelected>>", on_product_select)

    tk.Button(frame_sp, text="Thêm SP", command=lambda: add_item()).grid(row=0, column=6, padx=5)

    # ===========================================================
    #                     DANH SÁCH SẢN PHẨM
    # ===========================================================
    tree = ttk.Treeview(win, columns=("id","ten","sl","gia","tien"), show="headings", height=8)
    for c in ("id","ten","sl","gia","tien"):
        tree.heading(c, text=c.upper())
        tree.column(c, anchor="center", width=140)
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    def add_item():
        if not cbo_sp.get():
            messagebox.showwarning("Thiếu", "Hãy chọn sản phẩm")
            return
        id_sp = int(cbo_sp.get().split(" - ")[0])
        ten_sp = cbo_sp.get().split(" - ")[1]
        so_luong = int(sl.get())
        don_gia = float(gia_sp.get())
        thanh_tien = so_luong * don_gia
        tree.insert("", "end", values=(id_sp, ten_sp, so_luong, don_gia, thanh_tien))

    # ===========================================================
    #                       LƯU / IN HÓA ĐƠN
    # ===========================================================
    def save_invoice():
        if not cbo_nv.get() or not cbo_kh.get():
            messagebox.showwarning("Thiếu", "Hãy chọn nhân viên và khách hàng")
            return

        id_nv = int(cbo_nv.get().split(" - ")[0])
        id_kh = int(cbo_kh.get().split(" - ")[0])

        cn = connect_mysql()
        cur = cn.cursor()
        cur.execute("INSERT INTO hoadon(id_nv,id_khachhang,ngay_lap,tong_tien) VALUES (%s,%s,%s,%s)",
                    (id_nv, id_kh, ngay.get(), 0))
        cn.commit()

        id_hd = cur.lastrowid
        global last_invoice_id
        last_invoice_id = cur.lastrowid
        total = 0.0

        for r in tree.get_children():
            row = tree.item(r)["values"]
            id_sp = int(row[0])
            so_luong = int(row[2])
            don_gia = float(row[3])
            thanh_tien = float(row[4])

            cur.execute("INSERT INTO ct_hoadon(id_hoadon,id_sanpham,so_luong,don_gia,thanh_tien) "
                        "VALUES (%s,%s,%s,%s,%s)",
                        (id_hd, id_sp, so_luong, don_gia, thanh_tien))
            total += thanh_tien

        cur.execute("UPDATE hoadon SET tong_tien=%s WHERE id_hoadon=%s", (total, id_hd))
        cn.commit()

        messagebox.showinfo("OK", f"Đã lưu hóa đơn {id_hd}")
        
        # Bật nút in và gán ID vừa lưu
        btn_print.config(state="normal", command=lambda: print_invoice(id_hd))


    def print_invoice(invoice_id):
        try:
            cn = connect_mysql()
            cur = cn.cursor()
    
            # Lấy thông tin hóa đơn + chi tiết
            cur.execute("""
                SELECT hd.id_hoadon, hd.ngay_lap, hd.tong_tien,
                       nv.ho_ten, nv.sdt, nv.dia_chi,
                       kh.ho_ten, kh.sdt, kh.dia_chi,
                       sp.ten_sanpham, cthd.so_luong, cthd.don_gia, cthd.thanh_tien
                FROM hoadon hd
                JOIN nhanvien nv ON hd.id_nv = nv.id_nv
                JOIN khachhang kh ON hd.id_khachhang = kh.id_khachhang
                JOIN ct_hoadon cthd ON hd.id_hoadon = cthd.id_hoadon
                JOIN sanpham sp ON cthd.id_sanpham = sp.id_sanpham
                WHERE hd.id_hoadon = %s
            """, (invoice_id,))
            rows = cur.fetchall()
            if not rows:
                print("Không tìm thấy hóa đơn")
                return
    
            wb = Workbook()
            ws = wb.active
    
            # ========= Tiêu đề =========
            ws.merge_cells("A1:D1")
            ws["A1"] = "HÓA ĐƠN"
            ws["A1"].font = Font(size=20, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")
    
            # ========= Thông tin chung =========
            ws.append([])
            ws.append(["Mã HĐ:", rows[0][0]])
            ws.append(["Ngày lập:", rows[0][1]])
            ws.append(["Tổng tiền:", rows[0][2]])
            ws.append([])
            ws.append(["Nhân viên", "SĐT NV", "Địa chỉ NV", "Khách hàng", "SĐT KH", "Địa chỉ KH"])
            ws.append([rows[0][3], rows[0][4], rows[0][5], rows[0][6], rows[0][7], rows[0][8]])
    
            # ========= Bảng chi tiết sản phẩm =========
            ws.append([])
            ws.append(["Sản phẩm", "Số lượng", "Đơn giá", "Thành tiền"])
    
            # Border style
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
    
            start_row = ws.max_row + 1
            for r in rows:
                ws.append([r[9], r[10], r[11], r[12]])
    
            # Thêm border cho bảng
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                    min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
    
            wb.save(f"invoice_{invoice_id}.xlsx")
            print(f"Đã xuất hóa đơn ra invoice_{invoice_id}.xlsx")
    
        finally:
            cur.close()
            cn.close()

    # ---------------- NÚT LƯU + IN ----------------
    actions_frame = tk.Frame(win)
    actions_frame.pack(pady=10)

    tk.Button(actions_frame, text="Lưu hóa đơn", width=15, command=save_invoice)\
        .pack(side=tk.LEFT, padx=15)

    
   
    # Tạo nút in hóa đơn nhưng disable ban đầu
    btn_print = tk.Button(actions_frame, text="In hóa đơn", width=15, state="disabled")
    btn_print.pack(side=tk.LEFT, padx=15)



    actions_frame.pack(anchor="center")